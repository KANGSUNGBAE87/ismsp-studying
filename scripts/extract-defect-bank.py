from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


CODE_RE = re.compile(r"^(?:1|2|3)\.\d{1,2}\.\d{1,2}$")
GROUP_RE = re.compile(r"^(?:1|2|3)\.\d{1,2}\.?$")

# How many similar criteria to keep per criterion code for modes 2 and 3.
SIMILAR_LIST_SIZE = 12


def cell(value: Any) -> str:
  if value is None:
    return ""
  return re.sub(r"\s+", " ", str(value)).strip()


def normalize_name(path: Path) -> str:
  return unicodedata.normalize("NFC", path.name)


def normalize_text(value: Any) -> str:
  """Collapse whitespace and case for dedupe/comparison of Korean text."""
  return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(value or ""))).lower()


def char_ngrams(value: str, size: int = 2) -> set[str]:
  text = re.sub(r"\s+", "", value or "")
  if len(text) < size:
    return {text} if text else set()
  return {text[index : index + size] for index in range(len(text) - size + 1)}


def text_similarity(left: str, right: str) -> float:
  left_grams = char_ngrams(left)
  right_grams = char_ngrams(right)
  if not left_grams or not right_grams:
    return 0.0
  intersection = len(left_grams & right_grams)
  return (2 * intersection) / (len(left_grams) + len(right_grams))


def find_file(source_dir: Path, *parts: str) -> Path:
  for path in sorted(source_dir.iterdir(), key=lambda item: normalize_name(item)):
    name = normalize_name(path)
    if path.is_file() and all(part in name for part in parts):
      return path
  raise FileNotFoundError(f"Could not find file containing: {parts}")


def load_criteria(path: Path) -> dict[str, dict[str, str]]:
  workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
  criteria: dict[str, dict[str, str]] = {}

  for sheet in workbook.worksheets:
    group_code = ""
    group_name = ""
    current_code = ""
    for row in sheet.iter_rows(min_row=3, values_only=True):
      cells = [cell(value) for value in row]
      if len(cells) < 7:
        continue

      if GROUP_RE.match(cells[1]):
        group_code = cells[1]
        group_name = cells[2]

      if CODE_RE.match(cells[3]):
        current_code = cells[3]
        criteria[current_code] = {
          "code": cells[3],
          "name": cells[4],
          "domain": sheet.title,
          "groupCode": group_code,
          "groupName": group_name,
          "requirement": cells[5],
          "checkItemCount": 0,
        }

      if current_code and cells[6]:
        criteria[current_code]["checkItemCount"] += 1

  return criteria


def load_one_line_hints(path: Path) -> dict[str, dict[str, str]]:
  workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
  sheet = workbook.active
  hints: dict[str, dict[str, str]] = {}

  for row in sheet.iter_rows(values_only=True):
    cells = [cell(value) for value in row]
    if len(cells) < 3 or not CODE_RE.match(cells[0]):
      continue

    hints[cells[0]] = {
      "keyword": cells[2],
      "what": cells[3] if len(cells) > 3 else "",
      "who": cells[4] if len(cells) > 4 else "",
      "how": cells[5] if len(cells) > 5 else "",
      "when": cells[6] if len(cells) > 6 else "",
      "where": cells[7] if len(cells) > 7 else "",
    }

  return hints


def load_defect_pools(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
  workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
  sheet = workbook["(참고) 소스1"]
  true_pool: list[dict[str, Any]] = []
  false_pool: list[dict[str, Any]] = []

  for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    cells = [cell(value) for value in row]
    if len(cells) < 9:
      continue

    criteria_code = cells[3]
    criteria_name = cells[4]
    defect_case = cells[5]
    statement = cells[7] or f"{defect_case} {criteria_code} {criteria_name} 결함에 해당한다."
    truth = cells[8]

    if not CODE_RE.match(criteria_code) or truth not in {"0", "1"} or not statement:
      continue

    item = {
      "id": f"{'t' if truth == '1' else 'f'}-{row_index}",
      "statement": statement,
      "criteriaCode": criteria_code,
      "criteriaName": criteria_name,
      "defectCase": defect_case,
      "isCorrect": truth == "1",
      "sourceRow": row_index,
    }

    if truth == "1":
      true_pool.append(item)
    else:
      false_pool.append(item)

  return true_pool, false_pool


def build_defect_case_pool(true_pool: list[dict[str, Any]]) -> list[dict[str, Any]]:
  """One criterion-choice item per unique confirmed defect case."""
  pool: list[dict[str, Any]] = []
  seen: set[str] = set()

  for item in true_pool:
    key = normalize_text(item.get("defectCase"))
    if not key or key in seen:
      continue
    seen.add(key)
    pool.append(
      {
        "id": f"d-{len(pool) + 1}",
        "defectCase": item["defectCase"],
        "criteriaCode": item["criteriaCode"],
        "criteriaName": item["criteriaName"],
        "sourceRow": item.get("sourceRow"),
      }
    )

  return pool


def _read_confirmation_sheet(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
  sheet = workbook[sheet_name]
  rows: list[dict[str, Any]] = []
  group_code = ""
  group_name = ""

  for row_index, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
    cells = [cell(value) for value in row]
    if len(cells) < 6:
      continue

    if cells[1] and GROUP_RE.match(cells[1]):
      group_code = cells[1]
      group_name = cells[2]

    code = cells[3]
    question = cells[5]
    if not CODE_RE.match(code) or not question:
      continue

    rows.append(
      {
        "code": code,
        "name": cells[4],
        "question": question,
        "groupCode": group_code,
        "groupName": group_name,
        "keywords": cells[6] if len(cells) > 6 else "",
        "sourceRow": row_index,
      }
    )

  return rows


def load_check_item_pool(
  path: Path, criteria: dict[str, dict[str, str]]
) -> list[dict[str, Any]]:
  """Build the confirmation-item pool from the self-test workbook.

  `1.7 확인(통합)` is the main pool. Items whose normalized text appears in
  `1.3 확인(가상)` but not in `1.1 확인(일반)` are tagged as virtual-asset specific.
  """
  workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

  general_rows = _read_confirmation_sheet(workbook, "1.1 확인(일반)")
  virtual_rows = _read_confirmation_sheet(workbook, "1.3 확인(가상)")
  integrated_rows = _read_confirmation_sheet(workbook, "1.7 확인(통합)")

  general_set = {normalize_text(row["question"]) for row in general_rows}
  virtual_set = {normalize_text(row["question"]) for row in virtual_rows}
  virtual_specific = virtual_set - general_set

  pool: list[dict[str, Any]] = []
  seen: set[str] = set()

  for row in integrated_rows:
    key = normalize_text(row["question"])
    if not key or key in seen:
      continue
    seen.add(key)

    code = row["code"]
    is_virtual = key in virtual_specific
    pool.append(
      {
        "id": f"c-{len(pool) + 1}",
        "question": row["question"],
        "criteriaCode": code,
        "criteriaName": row["name"] or criteria.get(code, {}).get("name", ""),
        "groupCode": row["groupCode"] or criteria.get(code, {}).get("groupCode", ""),
        "groupName": row["groupName"] or criteria.get(code, {}).get("groupName", ""),
        "domain": criteria.get(code, {}).get("domain", ""),
        "sourceSheet": "1.7 확인(통합)",
        "sourceRow": row["sourceRow"],
        "sourceVariant": "virtualAsset" if is_virtual else "integrated",
        "displayBadge": "가상자산" if is_virtual else "",
        "keywords": row["keywords"],
      }
    )

  return pool


def build_similar_criteria_by_code(
  criteria: dict[str, dict[str, str]],
  true_pool: list[dict[str, Any]],
  false_pool: list[dict[str, Any]],
) -> dict[str, list[str]]:
  """Precompute a ranked list of similar criterion codes per criterion.

  Ranking signals (deterministic):
  - workbook false mappings for the same defect case (strongest)
  - same group prefix
  - same domain
  - name/requirement text similarity
  """
  # Map each defect case to its confirmed correct criterion.
  true_code_by_case: dict[str, str] = {}
  for item in true_pool:
    key = normalize_text(item.get("defectCase"))
    if key:
      true_code_by_case[key] = item["criteriaCode"]

  # For each correct code, count the wrong codes the workbook paired with the
  # same defect case. These are the most confusable distractors.
  strong_signal: dict[str, dict[str, int]] = {}
  for item in false_pool:
    correct = true_code_by_case.get(normalize_text(item.get("defectCase")))
    wrong = item["criteriaCode"]
    if not correct or correct not in criteria or wrong not in criteria or wrong == correct:
      continue
    strong_signal.setdefault(correct, {})
    strong_signal[correct][wrong] = strong_signal[correct].get(wrong, 0) + 1

  codes = list(criteria)
  similar: dict[str, list[str]] = {}

  for code in codes:
    base = criteria[code]
    base_text = f"{base.get('name', '')} {base.get('requirement', '')}"
    base_group = base.get("groupCode", "")
    base_domain = base.get("domain", "")
    code_strong = strong_signal.get(code, {})

    scored: list[tuple[str, float]] = []
    for other in codes:
      if other == code:
        continue
      candidate = criteria[other]
      score = 0.0
      if base_group and base_group == candidate.get("groupCode"):
        score += 0.6
      if base_domain and base_domain == candidate.get("domain"):
        score += 0.2
      score += 0.7 * text_similarity(
        base_text, f"{candidate.get('name', '')} {candidate.get('requirement', '')}"
      )
      # Strong workbook signal, capped so it cannot fully crowd the list.
      score += min(code_strong.get(other, 0), 5) * 0.5
      if score > 0:
        scored.append((other, score))

    scored.sort(key=lambda pair: (-pair[1], pair[0]))
    similar[code] = [other for other, _ in scored[:SIMILAR_LIST_SIZE]]

  return similar


def main() -> None:
  parser = argparse.ArgumentParser()
  parser.add_argument(
    "--source-dir",
    default="/Users/kangsungbae/Downloads/drive-download-20260613T123004Z-3-001",
  )
  parser.add_argument("--output", default="public/data/ismsp-defect-bank.json")
  args = parser.parse_args()

  source_dir = Path(args.source_dir)
  criteria_path = find_file(source_dir, "2023.10.31")
  defect_path = find_file(source_dir, "결함 테스트_240712v2")
  one_line_path = find_file(source_dir, "1줄 정리")
  self_test_path = find_file(source_dir, "셀프테스트")

  criteria = load_criteria(criteria_path)
  hints = load_one_line_hints(one_line_path)
  for code, hint in hints.items():
    if code in criteria:
      criteria[code]["hint"] = hint

  true_pool, false_pool = load_defect_pools(defect_path)
  defect_case_pool = build_defect_case_pool(true_pool)
  check_item_pool = load_check_item_pool(self_test_path, criteria)
  similar_criteria_by_code = build_similar_criteria_by_code(criteria, true_pool, false_pool)

  unknown_codes = sorted(
    {
      item["criteriaCode"]
      for item in [*true_pool, *false_pool, *check_item_pool]
      if item["criteriaCode"] not in criteria
    }
  )
  virtual_asset_count = sum(
    1 for item in check_item_pool if item["sourceVariant"] == "virtualAsset"
  )

  payload = {
    "version": "0.2.0",
    "generatedAt": datetime.now(timezone.utc).isoformat(),
    "sources": {
      "criteria": str(criteria_path),
      "defectTest": str(defect_path),
      "oneLineHints": str(one_line_path),
      "selfTest": str(self_test_path),
    },
    "counts": {
      "criteria": len(criteria),
      "truePool": len(true_pool),
      "falsePool": len(false_pool),
      "defectCasePool": len(defect_case_pool),
      "checkItemPool": len(check_item_pool),
      "virtualAssetCheckItems": virtual_asset_count,
      "unknownCodes": len(unknown_codes),
    },
    "unknownCodes": unknown_codes,
    "criteria": criteria,
    "similarCriteriaByCode": similar_criteria_by_code,
    "truePool": true_pool,
    "falsePool": false_pool,
    "defectCasePool": defect_case_pool,
    "checkItemPool": check_item_pool,
  }

  output = Path(args.output)
  output.parent.mkdir(parents=True, exist_ok=True)
  output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
  print(
    json.dumps(
      {
        "output": str(output),
        "criteria": len(criteria),
        "truePool": len(true_pool),
        "falsePool": len(false_pool),
        "defectCasePool": len(defect_case_pool),
        "checkItemPool": len(check_item_pool),
        "virtualAssetCheckItems": virtual_asset_count,
        "unknownCodes": len(unknown_codes),
      },
      ensure_ascii=False,
    )
  )


if __name__ == "__main__":
  main()
