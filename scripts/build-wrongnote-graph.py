from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any


CODE_RE = re.compile(r"^(?:1|2|3)\.\d{1,2}\.\d{1,2}$")
JUDGMENT_RE = re.compile(
  r"^(?P<caseText>.+?)\s+(?P<criteriaCode>\d\.\d+\.\d+)\s+(?P<criteriaName>.+?)\s+결함에\s+해당한다\.?$"
)

CONCEPTS: dict[str, dict[str, list[str]]] = {
  "asset": {
    "정보시스템": ["정보시스템", "서버", "DBMS", "네트워크"],
    "개인정보처리시스템": ["개인정보처리시스템"],
    "소스코드": ["소스코드", "source code", "프로그램 소스"],
    "설정파일": ["설정파일", "설정 파일", "config"],
    "비밀번호": ["비밀번호", "패스워드"],
    "인증키": ["인증키", "API Key", "토큰", "비밀키"],
    "로그": ["로그", "접속기록"],
    "월렛룸": ["월렛룸", "콜드월렛", "핫월렛"],
  },
  "control": {
    "경영진 보고": ["경영진", "보고", "의사결정"],
    "접근통제": ["접근통제", "접근권한", "접근 권한", "접근 제한"],
    "암호화": ["암호화", "암호정책", "암호 정책"],
    "암호키 관리": ["암호키", "키 관리"],
    "로그점검": ["로그점검", "접속기록 점검", "모니터링"],
    "백업": ["백업", "복구"],
    "파기": ["파기", "삭제"],
    "교육": ["교육", "훈련"],
  },
  "failure_mode": {
    "평문 저장": ["평문 저장", "평문으로 저장", "암호화하지 않고 저장"],
    "미수행": ["미수행", "수행하지 않은", "실시하지 않은"],
    "누락": ["누락", "포함하지 않은", "빠진"],
    "미확인": ["미확인", "확인되지 않은", "확인하지 않은"],
    "미검토": ["미검토", "검토하지 않은"],
    "미승인": ["미승인", "승인 없이", "승인 내역이 존재하지"],
    "기록 없음": ["기록 없음", "이력 없음", "증거자료가 확인되지 않은", "증적 없음"],
  },
  "context": {
    "가상자산": ["가상자산", "월렛", "핫월렛", "콜드월렛"],
    "개인정보": ["개인정보", "정보주체", "개인정보처리자"],
    "클라우드": ["클라우드", "cloud"],
    "위탁": ["위탁", "수탁자", "수탁사", "외부업체"],
  },
}


def normalize_text(value: Any) -> str:
  text = unicodedata.normalize("NFKC", str(value or "")).lower()
  text = re.sub(r"[\s\t\r\n]+", "", text)
  text = re.sub(r"[\p{P}\p{S}]+", "", text) if False else re.sub(r"[^\w가-힣]+", "", text)
  return text


def display_text(value: Any) -> str:
  return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value or ""))).strip()


def stable_hash(value: Any) -> str:
  return hashlib.sha1(normalize_text(value).encode("utf-8")).hexdigest()[:16]


def read_json(path: Path, default: Any = None) -> Any:
  if not path or not path.exists():
    return default
  return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
  path.parent.mkdir(parents=True, exist_ok=True)
  path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def node(node_id: str, node_type: str, label: str, properties: dict[str, Any] | None = None) -> dict[str, Any]:
  return {
    "id": node_id,
    "type": node_type,
    "label": label,
    "properties": properties or {},
  }


def edge(
  source: str,
  target: str,
  edge_type: str,
  properties: dict[str, Any] | None = None,
) -> dict[str, Any]:
  return {
    "source": source,
    "target": target,
    "type": edge_type,
    "properties": properties or {},
  }


class GraphBuilder:
  def __init__(self, bank: dict[str, Any], synth: list[dict[str, Any]], wrongnotes: dict[str, list[dict[str, Any]]]):
    self.bank = bank
    self.synth = synth
    self.wrongnotes = wrongnotes
    self.nodes: dict[str, dict[str, Any]] = {}
    self.edges: dict[tuple[str, str, str], dict[str, Any]] = {}
    self.criteria_summary: dict[str, dict[str, Any]] = {}
    self.match_report: dict[str, Any] = {
      "generatedAt": datetime.now(timezone.utc).isoformat(),
      "counts": defaultdict(int),
      "items": [],
    }
    self.check_index: dict[str, str] = {}
    self.defect_index: dict[str, str] = {}
    self.true_by_case: dict[str, dict[str, Any]] = {}

  def add_node(self, item: dict[str, Any]) -> None:
    existing = self.nodes.get(item["id"])
    if not existing:
      self.nodes[item["id"]] = item
      return
    existing["properties"] = {**existing.get("properties", {}), **item.get("properties", {})}

  def add_edge(self, item: dict[str, Any]) -> None:
    key = (item["source"], item["target"], item["type"])
    existing = self.edges.get(key)
    if not existing:
      props = dict(item.get("properties", {}))
      props.setdefault("count", 1)
      item["properties"] = props
      self.edges[key] = item
      return
    existing["properties"]["count"] = existing["properties"].get("count", 1) + item.get("properties", {}).get("count", 1)
    reasons = set(existing["properties"].get("reason", []))
    reasons.update(item.get("properties", {}).get("reason", []))
    if reasons:
      existing["properties"]["reason"] = sorted(reasons)

  def build(self) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    self._criteria()
    self._check_items()
    self._defect_cases()
    self._judgments()
    self._similar_edges()
    self._wrongnotes()
    self._finalize_summary()

    graph = {
      "version": "0.1.0",
      "generatedAt": datetime.now(timezone.utc).isoformat(),
      "nodes": sorted(self.nodes.values(), key=lambda item: item["id"]),
      "edges": sorted(self.edges.values(), key=lambda item: (item["source"], item["type"], item["target"])),
    }
    report = {
      **self.match_report,
      "counts": dict(self.match_report["counts"]),
    }
    summary = {
      "version": "0.1.0",
      "generatedAt": datetime.now(timezone.utc).isoformat(),
      "criteria": dict(sorted(self.criteria_summary.items())),
    }
    return graph, summary, report

  def _criteria(self) -> None:
    for code, item in (self.bank.get("criteria") or {}).items():
      criteria_id = criteria_node_id(code)
      domain = item.get("domain", "")
      domain_code = domain[:1] if domain else code[:1]
      group_code = normalize_group_code(item.get("groupCode", ""))
      domain_id = f"domain:{domain_code}"
      group_id = f"group:{group_code}"

      self.add_node(node(domain_id, "Domain", domain, {"domainCode": domain_code, "domainName": domain}))
      self.add_node(
        node(
          group_id,
          "CriteriaGroup",
          f"{group_code} {item.get('groupName', '')}".strip(),
          {
            "groupCode": group_code,
            "groupName": item.get("groupName", ""),
            "domainCode": domain_code,
          },
        )
      )
      self.add_node(
        node(
          criteria_id,
          "Criteria",
          f"{code} {item.get('name', '')}".strip(),
          {
            **item,
            "id": criteria_id,
            "code": code,
            "groupCode": group_code,
          },
        )
      )
      self.add_edge(edge(domain_id, group_id, "HAS_GROUP"))
      self.add_edge(edge(group_id, criteria_id, "HAS_CRITERIA"))
      self.add_edge(edge(criteria_id, group_id, "BELONGS_TO_GROUP"))
      self._summary_for(code, item)
      self._link_concepts(criteria_id, "Criteria", f"{item.get('name', '')} {item.get('requirement', '')}")

  def _check_items(self) -> None:
    for item in self.bank.get("checkItemPool") or []:
      text = item.get("question", "")
      item_id = f"checkitem:{item.get('id')}" if item.get("id") else check_node_id(text)
      code = item.get("criteriaCode", "")
      self.check_index[normalize_text(text)] = item_id
      self.add_node(
        node(
          item_id,
          "CheckItem",
          truncate_label(text),
          {
            **item,
        "id": item_id,
        "normalizedTextHash": stable_hash(text),
          },
        )
      )
      if code:
        self.add_edge(edge(item_id, criteria_node_id(code), "CHECKS"))
      self._link_concepts(item_id, "CheckItem", f"{text} {item.get('keywords', '')}")

  def _defect_cases(self) -> None:
    for item in self.bank.get("defectCasePool") or []:
      text = item.get("defectCase", "")
      item_id = defect_node_id(text)
      code = item.get("criteriaCode", "")
      self.defect_index[normalize_text(text)] = item_id
      self.add_node(
        node(
          item_id,
          "DefectCase",
          truncate_label(text),
          {
            **item,
            "id": item_id,
            "normalizedTextHash": stable_hash(text),
          },
        )
      )
      if code:
        self.add_edge(edge(item_id, criteria_node_id(code), "VIOLATES"))
      self._link_concepts(item_id, "DefectCase", text)

    for item in self.bank.get("truePool") or []:
      case_key = normalize_text(item.get("defectCase"))
      if case_key:
        self.true_by_case[case_key] = item

  def _judgments(self) -> None:
    for item in self.bank.get("truePool") or []:
      self._add_judgment(item, source="workbook", correct_code=item.get("criteriaCode"), correct_name=item.get("criteriaName"))

    for item in self.bank.get("falsePool") or []:
      true_item = self.true_by_case.get(normalize_text(item.get("defectCase")))
      correct_code = true_item.get("criteriaCode") if true_item else None
      correct_name = true_item.get("criteriaName") if true_item else None
      self._add_judgment(item, source="workbook", correct_code=correct_code, correct_name=correct_name)

    for item in self.synth:
      parsed = parse_judgment(item.get("statement", ""))
      claimed_code = clean_code(item.get("criteriaCode")) or clean_code(parsed.get("criteriaCode"))
      correct_code = (
        clean_code(item.get("correctAnswer"))
        or self.true_by_case.get(normalize_text(item.get("defectCase")), {}).get("criteriaCode")
      )
      claimed_criteria = (self.bank.get("criteria") or {}).get(claimed_code, {})
      correct_criteria = (self.bank.get("criteria") or {}).get(correct_code, {})
      normalized = {
        "id": item.get("id"),
        "statement": item.get("statement", ""),
        "defectCase": item.get("defectCase", ""),
        "criteriaCode": claimed_code,
        "criteriaName": item.get("criteriaName") or claimed_criteria.get("name", ""),
        "isCorrect": bool(item.get("isCorrect")),
        "family": item.get("family", ""),
        "trapAxis": item.get("trapAxis", ""),
        "difficulty": item.get("difficulty"),
        "wrongnessReason": item.get("wrongnessReason", ""),
        "sourceRow": item.get("sourceRow"),
      }
      self._add_judgment(
        normalized,
        source=f"synth:{item.get('source', 'unknown')}",
        correct_code=correct_code,
        correct_name=item.get("correctAnswerName") or correct_criteria.get("name", ""),
      )

  def _add_judgment(self, item: dict[str, Any], source: str, correct_code: str | None, correct_name: str | None) -> None:
    statement = item.get("statement", "")
    claimed_code = item.get("criteriaCode", "")
    is_correct = bool(item.get("isCorrect"))
    if not statement or not claimed_code:
      return

    parsed = parse_judgment(statement)
    case_text = item.get("defectCase") or parsed.get("caseText", "")
    judgment_key = f"{statement}|{source}|{item.get('id', '')}"
    judgment_id = f"judgment:{stable_hash(judgment_key)}"
    trap_type = "correct"
    if not is_correct:
      if correct_code and claimed_code != correct_code:
        trap_type = "criteria_misapplication"
      elif correct_code and claimed_code == correct_code:
        trap_type = "body_variant"
      else:
        trap_type = "orphan_false"

    self.add_node(
      node(
        judgment_id,
        "JudgmentStatement",
        truncate_label(statement),
        {
          "sourceId": item.get("id"),
          "statement": statement,
          "defectCase": case_text,
          "claimedCriteriaCode": claimed_code,
          "claimedCriteriaName": item.get("criteriaName", ""),
          "correctCriteriaCode": correct_code,
          "correctCriteriaName": correct_name,
          "isCorrect": is_correct,
          "trapType": trap_type,
          "family": item.get("family", ""),
          "trapAxis": item.get("trapAxis", ""),
          "difficulty": item.get("difficulty"),
          "wrongnessReason": item.get("wrongnessReason", ""),
          "source": source,
          "sourceRow": item.get("sourceRow"),
        },
      )
    )
    defect_id = self._resolve_defect_node(case_text, correct_code)
    if defect_id:
      self.add_edge(edge(judgment_id, defect_id, "ABOUT"))
    self.add_edge(edge(judgment_id, criteria_node_id(claimed_code), "CLAIMS_CRITERIA"))
    if correct_code:
      self.add_edge(edge(judgment_id, criteria_node_id(correct_code), "CORRECT_CRITERIA"))
    if trap_type == "criteria_misapplication" and correct_code:
      self.add_edge(
        edge(
          criteria_node_id(correct_code),
          criteria_node_id(claimed_code),
          "CONFUSED_WITH",
          {"reason": [source, item.get("family", "workbook") or "workbook"]},
        )
      )
    if trap_type == "body_variant":
      original_id = self._best_defect_for_code(case_text, correct_code)
      if original_id:
        self.add_edge(edge(judgment_id, original_id, "DEVIATES_FROM", {"reason": ["body_variant"]}))
    self._link_concepts(judgment_id, "JudgmentStatement", f"{case_text} {statement}")

  def _similar_edges(self) -> None:
    criteria = self.bank.get("criteria") or {}
    for code, similar_codes in (self.bank.get("similarCriteriaByCode") or {}).items():
      if code not in criteria:
        continue
      for rank, other in enumerate(similar_codes, start=1):
        if other in criteria and other != code:
          self.add_edge(
            edge(
              criteria_node_id(code),
              criteria_node_id(other),
              "SIMILAR_TO",
              {"rank": rank, "reason": ["precomputed_similar_criteria"]},
            )
          )

  def _wrongnotes(self) -> None:
    for note_type, pool_key, target_kind in [
      ("check_item", "checkItem", "CheckItem"),
      ("defect_case", "defectCase", "DefectCase"),
    ]:
      for raw in self.wrongnotes.get(pool_key, []):
        original_no = raw.get("no") or raw.get("originalNo") or raw.get("row")
        source_row = raw.get("row") or raw.get("sourceRow")
        text = display_text(raw.get("text") or raw.get("question") or raw.get("defectCase"))
        code = display_text(raw.get("criteriaCode") or raw.get("correctCriteriaCode"))
        name = display_text(raw.get("criteriaName") or raw.get("correctCriteriaName"))
        if original_no and (not text or not code):
          self.match_report["counts"]["blankNumberedRows"] += 1
          self.match_report["items"].append(
            {
              "noteType": note_type,
              "originalNo": original_no,
              "status": "blank_numbered_row",
            }
          )
          continue
        if not text or not code:
          continue

        wrongnote_id = f"wrongnote:{pool_key}:{source_row or original_no or stable_hash(text)}"
        match = self._match_wrongnote(text, code, target_kind)
        self.add_node(
          node(
            wrongnote_id,
            "WrongNoteEntry",
            truncate_label(text),
            {
              "noteType": note_type,
              "originalNo": original_no,
              "sourceRow": source_row,
              "text": text,
              "correctCriteriaCode": code,
              "correctCriteriaName": name,
              "matchType": match["matchType"],
              "matchedNodeId": match.get("matchedNodeId"),
              "similarityScore": match.get("similarityScore", 0),
              "result": "wrong",
              "sourceSheet": "확인사항오답" if note_type == "check_item" else "결함오답",
            },
          )
        )
        self.add_edge(edge(wrongnote_id, criteria_node_id(code), "WEAK_ON"))
        if match.get("matchedNodeId"):
          self.add_edge(
            edge(
              wrongnote_id,
              match["matchedNodeId"],
              "TARGETS",
              {"matchType": match["matchType"], "score": match.get("similarityScore", 0)},
            )
          )
        else:
          self.add_edge(edge(wrongnote_id, f"source:wrongnote:{pool_key}", "REVIEW_REQUIRED"))

        summary = self._summary_for(code)
        summary["wrongNoteCount"] += 1
        if note_type == "check_item":
          summary["checkItemWrongCount"] += 1
        else:
          summary["defectCaseWrongCount"] += 1
        self.match_report["counts"]["importedWrongNotes"] += 1
        self.match_report["counts"][match["matchType"]] += 1
        self.match_report["items"].append(
          {
            "noteType": note_type,
            "originalNo": original_no,
            "criteriaCode": code,
            "matchType": match["matchType"],
            "matchedNodeId": match.get("matchedNodeId"),
            "similarityScore": match.get("similarityScore", 0),
          }
        )
        self._link_concepts(wrongnote_id, "WrongNoteEntry", text)

  def _match_wrongnote(self, text: str, code: str, target_kind: str) -> dict[str, Any]:
    index = self.check_index if target_kind == "CheckItem" else self.defect_index
    key = normalize_text(text)
    if key in index:
      return {"matchType": "exact", "matchedNodeId": index[key], "similarityScore": 1.0}

    best_id = None
    best_score = 0.0
    target_nodes = [
      n
      for n in self.nodes.values()
      if n["type"] == target_kind and n.get("properties", {}).get("criteriaCode") == code
    ]
    if not target_nodes:
      target_nodes = [n for n in self.nodes.values() if n["type"] == target_kind]
    for candidate in target_nodes:
      candidate_text = candidate["properties"].get("question") or candidate["properties"].get("defectCase") or ""
      score = SequenceMatcher(None, normalize_text(text), normalize_text(candidate_text)).ratio()
      if score > best_score:
        best_id = candidate["id"]
        best_score = score

    if best_id and best_score >= 0.92:
      return {"matchType": "near", "matchedNodeId": best_id, "similarityScore": round(best_score, 4)}
    if best_id and best_score >= 0.85:
      return {"matchType": "review_required", "matchedNodeId": best_id, "similarityScore": round(best_score, 4)}
    return {"matchType": "standalone", "matchedNodeId": None, "similarityScore": round(best_score, 4)}

  def _resolve_defect_node(self, case_text: str, correct_code: str | None) -> str | None:
    key = normalize_text(case_text)
    if key in self.defect_index:
      return self.defect_index[key]
    return self._best_defect_for_code(case_text, correct_code)

  def _best_defect_for_code(self, case_text: str, correct_code: str | None) -> str | None:
    best_id = None
    best_score = 0.0
    for candidate in self.nodes.values():
      if candidate["type"] != "DefectCase":
        continue
      if correct_code and candidate["properties"].get("criteriaCode") != correct_code:
        continue
      score = SequenceMatcher(
        None,
        normalize_text(case_text),
        normalize_text(candidate["properties"].get("defectCase", "")),
      ).ratio()
      if score > best_score:
        best_id = candidate["id"]
        best_score = score
    return best_id if best_score >= 0.2 else None

  def _link_concepts(self, source_id: str, source_type: str, text: str) -> None:
    normalized = normalize_text(text)
    for concept_type, concepts in CONCEPTS.items():
      for name, aliases in concepts.items():
        if any(normalize_text(alias) in normalized for alias in aliases):
          concept_id = f"concept:{concept_type}:{stable_hash(name)}"
          self.add_node(
            node(
              concept_id,
              "Concept",
              name,
              {
                "conceptType": concept_type,
                "name": name,
                "aliases": aliases,
              },
            )
          )
          relation = {
            "asset": "AFFECTS_ASSET",
            "control": "HAS_CONTROL",
            "failure_mode": "HAS_FAILURE_MODE",
            "context": "HAS_CONTEXT",
          }.get(concept_type, "MENTIONS")
          self.add_edge(edge(source_id, concept_id, relation, {"sourceType": source_type}))
          self.add_edge(edge(source_id, concept_id, "MENTIONS", {"sourceType": source_type}))

  def _summary_for(self, code: str, criteria: dict[str, Any] | None = None) -> dict[str, Any]:
    if code not in self.criteria_summary:
      criteria = criteria or (self.bank.get("criteria") or {}).get(code, {})
      self.criteria_summary[code] = {
        "code": code,
        "name": criteria.get("name", ""),
        "domain": criteria.get("domain", ""),
        "groupCode": normalize_group_code(criteria.get("groupCode", "")),
        "groupName": criteria.get("groupName", ""),
        "wrongNoteCount": 0,
        "checkItemWrongCount": 0,
        "defectCaseWrongCount": 0,
        "judgmentWrongCount": 0,
        "confusionScore": 0,
        "similarCriteriaCount": 0,
        "weakScore": 0,
      }
    return self.criteria_summary[code]

  def _finalize_summary(self) -> None:
    for item in self.nodes.values():
      if item["type"] == "DefectCase":
        code = item["properties"].get("criteriaCode")
        if code in self.criteria_summary:
          self.criteria_summary[code]["defectCaseCount"] = self.criteria_summary[code].get("defectCaseCount", 0) + 1
      if item["type"] == "CheckItem":
        code = item["properties"].get("criteriaCode")
        if code in self.criteria_summary:
          self.criteria_summary[code]["checkItemCount"] = self.criteria_summary[code].get("checkItemCount", 0) + 1
    for item in self.edges.values():
      if item["type"] == "CONFUSED_WITH":
        code = item["source"].replace("criteria:", "")
        if code in self.criteria_summary:
          count = item["properties"].get("count", 1)
          self.criteria_summary[code]["confusionScore"] += count
          self.criteria_summary[code]["judgmentWrongCount"] += count
      if item["type"] == "SIMILAR_TO":
        code = item["source"].replace("criteria:", "")
        if code in self.criteria_summary:
          self.criteria_summary[code]["similarCriteriaCount"] += 1
    for code, item in self.criteria_summary.items():
      item["weakScore"] = (
        item.get("wrongNoteCount", 0) * 10
        + item.get("confusionScore", 0) * 2
        + item.get("defectCaseWrongCount", 0) * 3
        + item.get("checkItemWrongCount", 0) * 2
      )


def criteria_node_id(code: str) -> str:
  return f"criteria:{code}"


def check_node_id(text: str) -> str:
  return f"checkitem:{stable_hash(text)}"


def defect_node_id(text: str) -> str:
  return f"defect:{stable_hash(text)}"


def normalize_group_code(code: str) -> str:
  return display_text(code).rstrip(".")


def truncate_label(value: str, max_length: int = 80) -> str:
  text = display_text(value)
  return text if len(text) <= max_length else f"{text[: max_length - 1]}…"


def parse_judgment(statement: str) -> dict[str, str]:
  text = display_text(statement)
  match = JUDGMENT_RE.match(text)
  if match:
    return match.groupdict()
  code_match = re.search(r"(?P<criteriaCode>\d\.\d+\.\d+)\s*(?P<criteriaName>.*?)\s*결함에\s+해당한다\.?$", text)
  if not code_match:
    return {}
  return {
    "caseText": text[: code_match.start()].strip(),
    "criteriaCode": code_match.group("criteriaCode"),
    "criteriaName": code_match.group("criteriaName").strip(),
  }


def clean_code(value: Any) -> str:
  text = display_text(value)
  return text if CODE_RE.match(text) else ""


def load_wrongnotes_from_json(path: Path) -> dict[str, list[dict[str, Any]]]:
  data = read_json(path, {})
  return {
    "checkItem": data.get("checkItem", data.get("checkItems", [])),
    "defectCase": data.get("defectCase", data.get("defectCases", [])),
  }


def load_wrongnotes_from_xlsx(path: Path) -> dict[str, list[dict[str, Any]]]:
  try:
    import openpyxl
  except ImportError as exc:
    raise SystemExit("openpyxl is required to read xlsx wrong-note files.") from exc

  workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
  mapping = {
    "checkItem": "확인사항오답",
    "defectCase": "결함오답",
  }
  result: dict[str, list[dict[str, Any]]] = {"checkItem": [], "defectCase": []}
  for key, sheet_name in mapping.items():
    if sheet_name not in workbook.sheetnames:
      raise SystemExit(f"Wrong-note workbook is missing sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
      values = list(row)
      if not values or values[0] is None:
        continue
      result[key].append(
        {
          "row": row_index,
          "no": values[0],
          "text": values[1] if len(values) > 1 else "",
          "criteriaCode": values[2] if len(values) > 2 else "",
          "criteriaName": values[3] if len(values) > 3 else "",
        }
      )
  return result


def find_default_wrongnote(source_dir: Path) -> Path | None:
  if not source_dir.exists():
    return None
  for path in sorted(source_dir.iterdir(), key=lambda item: unicodedata.normalize("NFC", item.name)):
    name = unicodedata.normalize("NFC", path.name)
    if path.is_file() and name.endswith(".xlsx") and "오답노트" in name:
      return path
  return None


def main() -> None:
  parser = argparse.ArgumentParser()
  parser.add_argument("--bank", default="public/data/ismsp-defect-bank.json")
  parser.add_argument("--synth", default="public/data/synth-distractors.json")
  parser.add_argument("--wrongnote")
  parser.add_argument("--wrongnote-json")
  parser.add_argument("--source-dir", default="/Users/kangsungbae/Downloads/drive-download-20260613T123004Z-3-001")
  parser.add_argument("--graph-output", default="public/data/wrongnote-graph.json")
  parser.add_argument("--summary-output", default="public/data/criteria-summary.json")
  parser.add_argument("--report-output", default="public/data/wrongnote-match-report.json")
  args = parser.parse_args()

  bank = read_json(Path(args.bank), {})
  synth = read_json(Path(args.synth), [])
  if args.wrongnote_json:
    wrongnotes = load_wrongnotes_from_json(Path(args.wrongnote_json))
    wrongnote_path = Path(args.wrongnote_json)
  else:
    wrongnote_path = Path(args.wrongnote) if args.wrongnote else find_default_wrongnote(Path(args.source_dir))
    if not wrongnote_path:
      raise SystemExit("Could not find wrong-note workbook. Pass --wrongnote or --wrongnote-json.")
    wrongnotes = load_wrongnotes_from_xlsx(wrongnote_path)

  builder = GraphBuilder(bank, synth if isinstance(synth, list) else [], wrongnotes)
  graph, summary, report = builder.build()
  report["source"] = str(wrongnote_path)

  write_json(Path(args.graph_output), graph)
  write_json(Path(args.summary_output), summary)
  write_json(Path(args.report_output), report)
  print(
    json.dumps(
      {
        "graph": args.graph_output,
        "summary": args.summary_output,
        "report": args.report_output,
        "nodes": len(graph["nodes"]),
        "edges": len(graph["edges"]),
        "importedWrongNotes": report["counts"].get("importedWrongNotes", 0),
      },
      ensure_ascii=False,
    )
  )


if __name__ == "__main__":
  main()
