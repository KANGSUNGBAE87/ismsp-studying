from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from pypdf import PdfReader


CODE_RE = re.compile(r"\b(?:1|2|3)\.\d{1,2}\.\d{1,2}\b")
KEYWORDS = [
    "결함",
    "확인사항",
    "세부점검",
    "인증기준",
    "개인정보",
    "생명주기",
    "보호대책",
    "관리체계",
    "셀프테스트",
    "문제",
    "정답",
]


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def row_text(row: tuple[Any, ...]) -> str:
    return " | ".join(cell_to_text(v) for v in row if cell_to_text(v))


def inspect_xlsx(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    workbook_codes: Counter[str] = Counter()
    workbook_keywords: Counter[str] = Counter()

    for sheet in workbook.worksheets:
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        rows = []
        nonempty_rows = 0
        candidate_headers = []
        samples = []
        sheet_codes: Counter[str] = Counter()
        sheet_keywords: Counter[str] = Counter()

        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            text = row_text(row)
            if text:
                nonempty_rows += 1
                if len(samples) < 8:
                    samples.append({"row": idx, "text": text[:600]})
                if idx <= 20:
                    candidate_headers.append({"row": idx, "text": text[:600]})
                for code in CODE_RE.findall(text):
                    sheet_codes[code] += 1
                    workbook_codes[code] += 1
                for keyword in KEYWORDS:
                    if keyword in text:
                        sheet_keywords[keyword] += 1
                        workbook_keywords[keyword] += 1
            if idx <= 30:
                rows.append([cell_to_text(v) for v in row[: min(max_col, 12)]])

        sheets.append(
            {
                "name": sheet.title,
                "max_row": max_row,
                "max_col": max_col,
                "nonempty_rows": nonempty_rows,
                "candidate_headers": candidate_headers[:8],
                "samples": samples,
                "top_codes": sheet_codes.most_common(20),
                "keyword_counts": dict(sheet_keywords),
                "first_30_rows": rows,
            }
        )

    return {
        "type": "xlsx",
        "path": str(path),
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "sheet_count": len(workbook.sheetnames),
        "sheets": sheets,
        "top_codes": workbook_codes.most_common(30),
        "keyword_counts": dict(workbook_keywords),
    }


def inspect_pdf(path: Path) -> dict[str, Any]:
    reader = PdfReader(str(path))
    page_count = len(reader.pages)
    combined = []
    page_summaries = []
    code_counter: Counter[str] = Counter()
    keyword_counter: Counter[str] = Counter()

    for page_index, page in enumerate(reader.pages):
        try:
            text = page.extract_text() or ""
        except Exception as exc:  # pragma: no cover - best effort extraction
            text = f"[extract_error: {exc}]"
        text = re.sub(r"\s+", " ", text).strip()
        if page_index < 6:
            page_summaries.append({"page": page_index + 1, "text": text[:1200]})
        combined.append(text[:4000])
        for code in CODE_RE.findall(text):
            code_counter[code] += 1
        for keyword in KEYWORDS:
            if keyword in text:
                keyword_counter[keyword] += 1

    full_sample = " ".join(combined[:20])
    return {
        "type": "pdf",
        "path": str(path),
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "page_count": page_count,
        "page_summaries": page_summaries,
        "top_codes": code_counter.most_common(30),
        "keyword_counts": dict(keyword_counter),
        "sample": full_sample[:5000],
    }


def inspect_file(path: Path) -> dict[str, Any]:
    lower_name = path.name.lower()
    if lower_name.endswith(".xlsx"):
        return inspect_xlsx(path)
    if lower_name.endswith(".pdf") or ".pdf" in lower_name:
        return inspect_pdf(path)
    return {
        "type": "unknown",
        "path": str(path),
        "name": path.name,
        "size_bytes": path.stat().st_size,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    files = sorted([p for p in source_dir.iterdir() if p.is_file()], key=lambda p: p.name)
    result = {"source_dir": str(source_dir), "file_count": len(files), "files": []}

    for path in files:
        try:
            result["files"].append(inspect_file(path))
        except Exception as exc:
            result["files"].append(
                {
                    "type": "error",
                    "path": str(path),
                    "name": path.name,
                    "size_bytes": path.stat().st_size if path.exists() else None,
                    "error": repr(exc),
                }
            )

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(output), "file_count": len(files)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
