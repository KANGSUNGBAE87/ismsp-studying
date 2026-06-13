from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl


CODE_RE = re.compile(r"^(?:1|2|3)\.\d{1,2}\.\d{1,2}$")
GROUP_RE = re.compile(r"^(?:1|2|3)\.\d{1,2}\.?$")


def norm_name(path: Path) -> str:
    return unicodedata.normalize("NFC", path.name)


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def find_file(source_dir: Path, *contains: str) -> Path:
    normalized = [(norm_name(path), path) for path in source_dir.iterdir() if path.is_file()]
    for name, path in normalized:
        if all(part in name for part in contains):
            return path
    raise FileNotFoundError(contains)


def official_criteria(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    criteria: dict[str, dict[str, Any]] = {}
    check_items = []
    domain_counts = Counter()

    for sheet in workbook.worksheets:
        current_group = ""
        current_group_name = ""
        current_code = ""
        current_name = ""
        current_desc = ""
        for row in sheet.iter_rows(min_row=3, values_only=True):
            cells = [text(v) for v in row]
            if len(cells) < 7:
                continue
            if GROUP_RE.match(cells[1]):
                current_group = cells[1]
                current_group_name = cells[2]
            if CODE_RE.match(cells[3]):
                current_code = cells[3]
                current_name = cells[4]
                current_desc = cells[5]
                criteria.setdefault(
                    current_code,
                    {
                        "code": current_code,
                        "name": current_name,
                        "group": current_group,
                        "group_name": current_group_name,
                        "domain": sheet.title,
                        "description": current_desc,
                        "check_item_count": 0,
                    },
                )
                domain_counts[sheet.title] += 1
            check_item = cells[6]
            if current_code and check_item:
                criteria[current_code]["check_item_count"] += 1
                check_items.append(
                    {
                        "code": current_code,
                        "name": current_name,
                        "group": current_group,
                        "group_name": current_group_name,
                        "domain": sheet.title,
                        "check_item": check_item,
                    }
                )

    return {
        "file": path.name,
        "criteria_count": len(criteria),
        "check_item_count": len(check_items),
        "domain_counts": dict(domain_counts),
        "criteria": criteria,
        "check_items_sample": check_items[:10],
    }


def count_official_variant(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    criteria = set()
    check_count = 0
    sheet_counts = {}

    for sheet in workbook.worksheets:
        current_code = ""
        sheet_criteria = set()
        sheet_checks = 0
        for row in sheet.iter_rows(min_row=3, values_only=True):
            cells = [text(v) for v in row]
            if len(cells) < 7:
                continue
            if len(cells) > 3 and CODE_RE.match(cells[3]):
                current_code = cells[3]
                criteria.add(current_code)
                sheet_criteria.add(current_code)
            if current_code and len(cells) > 6 and cells[6]:
                check_count += 1
                sheet_checks += 1
        sheet_counts[sheet.title] = {"criteria": len(sheet_criteria), "check_items": sheet_checks}

    return {
        "file": path.name,
        "criteria_count": len(criteria),
        "check_item_count": check_count,
        "sheet_counts": sheet_counts,
    }


def selftest_summary(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    summary: dict[str, Any] = {"file": path.name, "sheets": {}, "code_counts": {}, "samples": {}}
    code_counts: Counter[str] = Counter()

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        sheet_name = sheet.title
        kind = "other"
        if "확인" in sheet_name and "답안지" not in sheet_name:
            kind = "check_source"
        elif "문제" in sheet_name:
            kind = "question_sample"
        elif "답안지" in sheet_name:
            kind = "answer_sheet"
        elif "결함" in sheet_name:
            kind = "defect_source"

        row_count = 0
        codes = Counter()
        samples = []
        for row in rows:
            cells = [text(v) for v in row]
            if not any(cells):
                continue
            row_count += 1
            row_codes = [c for c in cells if CODE_RE.match(c)]
            for code in row_codes:
                codes[code] += 1
                code_counts[code] += 1
            if len(samples) < 4:
                samples.append(" | ".join(c for c in cells if c)[:500])

        summary["sheets"][sheet_name] = {
            "kind": kind,
            "nonempty_rows": row_count,
            "unique_codes": len(codes),
            "top_codes": codes.most_common(10),
        }
        summary["samples"][sheet_name] = samples

    summary["code_counts"] = dict(code_counts)
    return summary


def defect_test_summary(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    source = workbook["(참고) 소스1"]
    true_by_code = Counter()
    false_by_code = Counter()
    unique_cases = set()
    true_pairs = set()
    samples = []

    for row in source.iter_rows(min_row=2, values_only=True):
        cells = [text(v) for v in row]
        if len(cells) < 9:
            continue
        code = cells[3]
        name = cells[4]
        defect_case = cells[5]
        truth = cells[8]
        if not CODE_RE.match(code) or not defect_case:
            continue
        unique_cases.add(defect_case)
        if truth == "1":
            true_by_code[code] += 1
            true_pairs.add((defect_case, code, name))
            if len(samples) < 8:
                samples.append({"case": defect_case[:300], "code": code, "name": name, "truth": truth})
        else:
            false_by_code[code] += 1

    question_sheet = workbook["문제 출제"]
    option_count = 0
    generated_question_count = 0
    for row in question_sheet.iter_rows(values_only=True):
        cells = [text(v) for v in row]
        if cells and re.match(r"^\d+\)", cells[0]):
            option_count += 1
        if cells and "ISMS-P 인증심사" in cells[0]:
            generated_question_count += 1

    return {
        "file": path.name,
        "source1_rows_true": sum(true_by_code.values()),
        "source1_rows_false": sum(false_by_code.values()),
        "unique_base_defect_cases": len(unique_cases),
        "true_defect_case_code_pairs": len(true_pairs),
        "true_code_counts": dict(true_by_code),
        "false_code_counts": dict(false_by_code),
        "top_true_codes": true_by_code.most_common(20),
        "top_false_codes": false_by_code.most_common(20),
        "generated_question_prompts": generated_question_count,
        "generated_options": option_count,
        "samples": samples,
    }


def wrong_note_summary(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[str, Any] = {"file": path.name, "sheets": {}}
    for sheet in workbook.worksheets:
        codes = Counter()
        rows = 0
        samples = []
        for row in sheet.iter_rows(values_only=True):
            cells = [text(v) for v in row]
            if not any(cells):
                continue
            rows += 1
            for cell in cells:
                if CODE_RE.match(cell):
                    codes[cell] += 1
            if len(samples) < 5:
                samples.append(" | ".join(c for c in cells if c)[:500])
        result["sheets"][sheet.title] = {
            "nonempty_rows": rows,
            "unique_codes": len(codes),
            "code_counts": dict(codes),
            "top_codes": codes.most_common(10),
            "samples": samples,
        }
    return result


def one_line_summary(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = 0
    codes = Counter()
    samples = []
    for row in sheet.iter_rows(values_only=True):
        cells = [text(v) for v in row]
        if not any(cells):
            continue
        rows += 1
        for cell in cells:
            if CODE_RE.match(cell):
                codes[cell] += 1
        if len(samples) < 8:
            samples.append(" | ".join(c for c in cells if c)[:500])
    return {
        "file": path.name,
        "nonempty_rows": rows,
        "unique_codes": len(codes),
        "code_counts": dict(codes),
        "top_codes": codes.most_common(20),
        "samples": samples,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    official = official_criteria(find_file(source_dir, "2023.10.31"))
    variants = [
        count_official_variant(find_file(source_dir, "(7의2)")),
        count_official_variant(find_file(source_dir, "(7의3)")),
    ]
    selftest = selftest_summary(find_file(source_dir, "셀프테스트_20240103"))
    defect_test = defect_test_summary(find_file(source_dir, "결함 테스트_240712v2"))
    wrong_note = wrong_note_summary(find_file(source_dir, "오답노트"))
    one_line = one_line_summary(find_file(source_dir, "1줄 정리"))

    official_codes = set(official["criteria"].keys())
    linked_code_sets = {
        "selftest": set(selftest["code_counts"].keys()),
        "defect_test_true": set(defect_test["true_code_counts"].keys()),
        "wrong_note": {
            code
            for sheet in wrong_note["sheets"].values()
            for code in sheet["code_counts"].keys()
        },
        "one_line": set(one_line["code_counts"].keys()),
    }

    result = {
        "official_basis": official,
        "official_variants": variants,
        "selftest": selftest,
        "defect_test": defect_test,
        "wrong_note": wrong_note,
        "one_line": one_line,
        "coverage": {
            name: {
                "codes_in_sample": len(codes),
                "codes_overlapping_official": len(codes & official_codes),
                "unknown_codes": sorted(codes - official_codes)[:30],
            }
            for name, codes in linked_code_sets.items()
        },
    }

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
